import os
import re

from django import shortcuts, urls
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import views as views_auth
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.forms import modelformset_factory
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    FormView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)

import rrgg.models
from rrgg import mixins as rrgg_mixins

from . import forms
from .utils import SeguroItem


class LoginView(views_auth.LoginView):
    template_name = "rrggweb/login.html"
    form_class = forms.LoginAuthenticationForm

    def form_valid(self, form):
        registrar_id = (
            form.get_user().consultant_membership.first().consultant.id
        )
        self.next_page = urls.reverse(
            "rrggweb:home", kwargs={"registrar_id": registrar_id}
        )
        return super().form_valid(form)


class LogoutView(views_auth.LogoutView):
    next_page = urls.reverse_lazy("rrggweb:login")


# SELLER


class SelectRoleSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/seller_form.html"
    form_class = forms.RoleForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Seleccionar asesor"
        return context


class SelectSellerSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/seller_form.html"
    form_class = forms.SellerForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        role_id = self.kwargs.get("role_id")
        kwargs["role_id"] = role_id
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Seleccionar asesor"
        context["role_selector"] = forms.RoleForm(
            role_id=self.kwargs["role_id"]
        )
        return context


class UpdateSellerSupportView(
    rrgg_mixins.RrggBootstrapDisplayMixin, LoginRequiredMixin, UpdateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/seller_form.html"
    model = rrgg.models.QuotationInsuranceVehicle
    fields = ["consultant_seller"]
    pk_url_kwarg = "quotation_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Cambiar asesor"
        return context


# CUSTOMER


class SearchCustomerSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/person_form.html"
    form_class = forms.SearchPersonForm

    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        if not re.match("^[0-9]+$", document_number):
            form.add_error(
                "document_number",
                "El número de documento debe contener solo números.",
            )
            return super().form_invalid(form)
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Buscar contratante"
        context["body"] = "Buscar contratante"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        return context


class SelectCustomerSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/person_form.html"
    form_class = forms.SelectCustomerForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Definir contratante"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        return context


class PersonFormSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin
):
    template_name = "rrggweb/quotation/insurance/vehicle/person_form.html"
    fields = "__all__"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["email2"].required = False
        form.fields["address"].required = False
        return form


class CreateNaturalPersonSupportView(PersonFormSupportView, CreateView):
    model = rrgg.models.NaturalPerson

    def get_initial(self):
        initial = super().get_initial()
        dni = rrgg.models.DocumentType.objects.get(code="dni")
        initial["document_type"] = dni
        initial["document_number"] = self.request.GET.get(
            "document_number", ""
        )
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["second_surname"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar contratante"
        context["body"] = "Formulario del Contratante"
        context["type_customer"] = "Persona natural"
        return context


class UpdateNaturalPersonSupportView(PersonFormSupportView, UpdateView):
    model = rrgg.models.NaturalPerson
    pk_url_kwarg = "natural_person_id"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["second_surname"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar contratante"
        context["body"] = "Formulario del Contratante"
        context["type_customer"] = "Persona natural"
        return context


class CreateLegalPersonSupportView(PersonFormSupportView, CreateView):
    model = rrgg.models.LegalPerson

    def get_initial(self):
        initial = super().get_initial()
        ruc = rrgg.models.DocumentType.objects.get(code="ruc")
        initial["document_type"] = ruc
        initial["document_number"] = self.request.GET.get(
            "document_number", ""
        )
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["general_manager"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar contratante"
        context["body"] = "Formulario del Contratante"
        context["type_customer"] = "Persona jurídica"
        return context


class UpdateLegalPersonSupportView(PersonFormSupportView, UpdateView):
    model = rrgg.models.LegalPerson
    pk_url_kwarg = "legal_person_id"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["general_manager"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar contratante"
        context["body"] = "Formulario del Contratante"
        context["type_customer"] = "Persona jurídica"
        return context


# VEHICLE


class SearchVehicleSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/vehicle_form.html"
    form_class = forms.SearchVehicleForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Buscar vehículo"
        context["body"] = "Buscar vehículo"
        context["pretty_style"] = False
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        return context


class CreateVehicleSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/vehicle_form.html"
    model = rrgg.models.Vehicle
    fields = "__all__"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["plate"].required = False
        form.fields["has_gps"].required = False
        form.fields["has_endorsee"].required = False
        form.fields["endorsement_bank"].required = False
        form.fields["class_type"].required = False
        return form

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:define_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_initial(self):
        initial = super().get_initial()
        initial["plate"] = self.request.GET.get("plate", "")
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar vehículo"
        context["pretty_style"] = True
        context["body"] = "Formulario del vehículo"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        return context


class UpdateVehicleSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/vehicle_form.html"
    model = rrgg.models.Vehicle
    fields = "__all__"
    pk_url_kwarg = "vehicle_id"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["class_type"].required = False
        form.fields["has_gps"].required = False
        form.fields["has_endorsee"].required = False
        form.fields["endorsement_bank"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar vehículo"
        context["body"] = "Formulario del vehículo"
        context["pretty_style"] = True
        return context


# OWNER


class DefineOwnerSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/person_form.html"
    form_class = forms.DefineOwnerForm

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["is_owner"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Definir asegurado"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        return context


class SearchOwnerSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/person_form.html"
    form_class = forms.SearchPersonForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Buscar asegurado"
        context["body"] = "Buscar asegurado"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        return context


class CreateOwnerSupportView(PersonFormSupportView, CreateView):
    model = rrgg.models.NaturalPerson

    def get_initial(self):
        initial = super().get_initial()
        dni = rrgg.models.DocumentType.objects.get(code="dni")
        initial["document_type"] = dni
        initial["document_number"] = self.request.GET.get(
            "document_number", ""
        )
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["second_surname"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar asegurado"
        context["body"] = "Formulario del Asegurado"
        context["type_customer"] = "Persona natural"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        return context


class UpdateOwnerSupportView(PersonFormSupportView, UpdateView):
    model = rrgg.models.NaturalPerson
    pk_url_kwarg = "owner_id"

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["second_surname"].required = False
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar asegurado"
        context["body"] = "Formulario del Asegurado"
        return context


# QUOTATION


class QuotationListSupportView(LoginRequiredMixin, ListView):
    model = rrgg.models.QuotationInsuranceVehicle
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get("q")
        if query:
            return rrgg.models.QuotationInsuranceVehicle.objects.filter(
                Q(source="quotation"),
                Q(customer__natural_person__given_name__icontains=query)
                | Q(
                    customer__natural_person__first_surname__icontains=query  # noqa: E501
                )
                | Q(
                    customer__natural_person__second_surname__icontains=query  # noqa: E501
                )
                | Q(
                    customer__natural_person__document_number__icontains=query  # noqa: E501
                )
                | Q(
                    customer__legal_person__registered_name__icontains=query  # noqa: E501
                )
                | Q(
                    customer__legal_person__general_manager__icontains=query  # noqa: E501
                )
                | Q(
                    customer__legal_person__document_number__icontains=query  # noqa: E501
                )
                | Q(consultant_seller__given_name__icontains=query)
                | Q(consultant_seller__first_surname__icontains=query)
                | Q(vehicle__plate__icontains=query),
            ).order_by("-id")
        else:
            return (
                super()
                .get_queryset()
                .filter(source="quotation")
                .order_by("-id")
            )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Lista de cotizaciones"
        context["new_register"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["search_query"] = self.request.GET.get("q", "")
        return context


class CreateQuotationSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/form.html"
    model = rrgg.models.QuotationInsuranceVehicle
    fields = ["insured_amount", "currency"]

    def form_valid(self, form):
        form.instance.risk_id = 1
        form.instance.consultant_registrar_id = self.kwargs["registrar_id"]
        form.instance.consultant_seller_id = self.kwargs["seller_id"]
        form.instance.customer_id = self.kwargs["customer_id"]
        form.instance.vehicle_id = self.kwargs["vehicle_id"]
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create_premiums",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        context["owner"] = context["vehicle"].ownership
        return context


class UpdateQuotationSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/form.html"
    model = rrgg.models.QuotationInsuranceVehicle
    fields = ["insured_amount", "currency"]
    pk_url_kwarg = "quotation_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar suma asegurada"
        return context


class DetailQuotationSupportView(LoginRequiredMixin, DetailView):
    template_name = "rrggweb/quotation/insurance/vehicle/detail.html"
    model = rrgg.models.QuotationInsuranceVehicle
    pk_url_kwarg = "quotation_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["customer"] = self.object.customer
        context["vehicle"] = self.object.vehicle
        context["owner"] = context["vehicle"].ownership
        return context


# PREMIUM


class PremiumListSupportView(LoginRequiredMixin, ListView):
    model = rrgg.models.QuotationInsuranceVehiclePremium
    paginate_by = 10

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Lista de primas vehiculares"
        return context


class CreatePremiumSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/premium_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium
    fields = ["insurance_vehicle_ratio", "amount", "rate"]

    def form_valid(self, form):
        ivr = shortcuts.get_object_or_404(
            rrgg.models.InsuranceVehicleRatio,
            id=form.cleaned_data["insurance_vehicle_ratio"].id,
        )
        form.instance.tax_percentage = ivr.tax
        form.instance.emission_right_percentage = ivr.emission_right
        form.instance.quotation_insurance_vehicle_id = self.kwargs[
            "quotation_id"
        ]
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar prima"
        context["body"] = "Formulario prima vehicular"
        return context


class UpdatePremiumSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/premium_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium
    fields = ["insurance_vehicle_ratio", "amount", "rate"]
    pk_url_kwarg = "premium_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar Prima de Aseguradora"
        context["insured_amount"] = (
            self.object.quotation_insurance_vehicle.insured_amount
        )
        return context


class DeletePremiumSupportView(LoginRequiredMixin, DeleteView):
    template_name = "rrggweb/quotation/insurance/vehicle/delete_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Eliminar Prima de Aseguradora"
        return context


# PREMIUM - QUOTATION - MULTIPLE SALE


class CreatePremiumQuotationSupportView(LoginRequiredMixin, CreateView):
    template_name = "rrggweb/issuance/insurance/vehicle/premium_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium
    form_class = forms.PremiumQuotationForm

    def form_valid(self, form):
        quotation = rrgg.models.QuotationInsuranceVehicle(
            risk_id=1,
            consultant_registrar_id=self.kwargs["seller_id"],
            consultant_seller_id=self.kwargs["seller_id"],
            customer_id=self.kwargs["customer_id"],
            currency_id=1,
            source="new_sale",
            vehicle_id=self.kwargs["vehicle_id"],
            insured_amount=form.cleaned_data["insured_amount"],
        )
        quotation.save()
        form.instance.amount = form.cleaned_data["amount"]
        form.instance.rate = form.cleaned_data["rate"]
        form.instance.insurance_vehicle_ratio_id = 1
        form.instance.quotation_insurance_vehicle_id = quotation.id
        form.instance.in_progress = True
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Registrar Prima vehicular"
        return context


class UpdatePremiumQuotationSupportView(LoginRequiredMixin, UpdateView):
    template_name = "rrggweb/issuance/insurance/vehicle/premium_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium
    form_class = forms.PremiumQuotationForm

    def get_object(self, queryset=None):
        return rrgg.models.QuotationInsuranceVehiclePremium.objects.get(
            pk=self.kwargs["premium_id"]
        )

    def _get_data(self):
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=premium.quotation_insurance_vehicle.id,
        )
        return premium, quotation

    def get_initial(self):
        initial = super().get_initial()
        initial["insured_amount"] = self._get_data()[1].insured_amount
        return initial

    def form_valid(self, form):
        quotation = rrgg.models.QuotationInsuranceVehicle.objects.get(
            pk=self.object.quotation_insurance_vehicle.id
        )
        quotation.insured_amount = form.cleaned_data["insured_amount"]
        quotation.save()
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": (
                    self.object.quotation_insurance_vehicle.consultant_seller.id  # noqa: E501
                ),
                "customer_id": (
                    self.object.quotation_insurance_vehicle.customer.id
                ),
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar Prima vehicular"
        return context


class DeletePremiumQuotationSupportView(LoginRequiredMixin, DeleteView):
    template_name = "rrggweb/issuance/insurance/vehicle/delete_premium.html"
    model = rrgg.models.QuotationInsuranceVehicle

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.premiums.first().delete()
            self.object.delete()
            return redirect(self.get_success_url())
        except ProtectedError:
            messages.error(
                request, "No se puede eliminar este registro vehicular"
            )
            return redirect(request.get_full_path())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Eliminar registro vehicular"
        return context


class SelectCurrencyInsuranceSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.CurrencyInsuranceForm

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Seleccionar compañía y moneda"
        return context


# PLAN


class SelectPlanSupportView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.InsurancePlanForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        qivp = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        ivr = qivp.insurance_vehicle_ratio
        riv = shortcuts.get_object_or_404(
            rrgg.models.RiskInsuranceVehicle,
            insurance_vehicle_id=ivr.insurance_vehicle_id,
            risk_id=qivp.quotation_insurance_vehicle.risk_id,
        )
        kwargs["riv_id"] = riv.id
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Seleccionar plan de seguro"
        context["premium"] = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        context["quotation"] = context["premium"].quotation_insurance_vehicle
        context["customer"] = context["quotation"].customer
        context["vehicle"] = context["quotation"].vehicle
        context["owner"] = context["vehicle"].ownership
        riv = shortcuts.get_object_or_404(
            rrgg.models.RiskInsuranceVehicle,
            id=self.get_form_kwargs()["riv_id"],
        )
        context["risk_selector"] = forms.RiskForm(risk_id=riv.risk_id)
        context["insurance_vehicle_selector"] = forms.InsuranceVehicleForm(
            iv_id=riv.insurance_vehicle_id
        )
        return context


# ISSUANCE


class IssuanceListSupportView(LoginRequiredMixin, ListView):
    template_name = "rrggweb/issuance/insurance/vehicle/list.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get("q")
        if query:
            return rrgg.models.IssuanceInsuranceVehicle.objects.filter(
                Q(policy__icontains=query)
                | Q(collection_document__icontains=query)
                | Q(issuance_date__icontains=query)
                | Q(issuance_type__name__icontains=query)
                | Q(initial_validity__icontains=query)
                | Q(final_validity__icontains=query)
                | Q(payment_method__name__icontains=query)
                | Q(consultant_seller__given_name__icontains=query)
                | Q(consultant_seller__first_surname__icontains=query)
                | Q(status__name__icontains=query)
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__natural_person__given_name__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__natural_person__first_surname__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__natural_person__second_surname__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__natural_person__document_number__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__legal_person__registered_name__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__legal_person__general_manager__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__customer__legal_person__document_number__icontains=query  # noqa: E501
                )
                | Q(
                    quotation_vehicle_premiums__quotation_insurance_vehicle__vehicle__plate__icontains=query  # noqa: E501
                )
            ).order_by("-id")
        else:
            return super().get_queryset().order_by("-id")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["subtitle"] = "Lista de emisiones"
        return context


class CreateIssuanceSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/form.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    fields = [
        "policy",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "payment_method",
    ]

    def get_initial(self):
        initial = super().get_initial()
        plan = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self.kwargs["plan_id"]
        )
        kcs_commission_percentage = round(plan.commission * 100, 1)
        initial["plan_commission_percentage"] = kcs_commission_percentage
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Crear emisión"
        context["premium"] = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )

        context["ratio"] = context["premium"].insurance_vehicle_ratio
        context["insurance_plan"] = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self.kwargs["plan_id"]
        )
        context["quotation"] = context["premium"].quotation_insurance_vehicle
        context["customer"] = context["quotation"].customer.pick
        context["vehicle"] = context["quotation"].vehicle
        if isinstance(
            context["vehicle"].ownership.pick, rrgg.models.NaturalPerson
        ):
            context["owner"] = context["vehicle"].ownership.pick
        return context


class UpdateIssuanceSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/form.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    fields = [
        "policy",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "payment_method",
    ]
    pk_url_kwarg = "issuance_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Editar Emisión"
        return context


# DOCUMENT


class GetDocumentView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        document_id = self.kwargs["document_id"]
        instance = get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicleDocument, pk=document_id
        )
        file_name = instance.file.name
        response = FileResponse(instance.file, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{file_name}"'
        return response


class AddDocumentSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/document_form.html"
    model = rrgg.models.IssuanceInsuranceVehicleDocument
    fields = ["issuance", "file"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Subir documentos"
        context["type"] = "create"
        context["documents"] = self.model.objects.filter(
            issuance_id=self.kwargs["issuance_id"]
        )
        context["finish"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class DeleteDocumentSupportView(LoginRequiredMixin, DeleteView):
    template_name = "rrggweb/issuance/insurance/vehicle/document_form.html"
    model = rrgg.models.IssuanceInsuranceVehicleDocument
    pk_url_kwarg = "document_id"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        file_path = os.path.join(settings.MEDIA_ROOT, self.object.file.name)
        os.remove(file_path)
        self.object.delete()
        return redirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Eliminar documento"
        context["type"] = "delete"
        return context


# ENDORSEMENTS


class CreateEndorsementSupportView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/endorsement_form.html"
    model = rrgg.models.EndorsementVehicle
    fields = [
        "insured_amount",
        "net_premium",
        "rate",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "currency",
        "payment_method",
        "detail",
    ]

    def _get_data(self):
        issuance = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        ivr = premium.insurance_vehicle_ratio
        vehicle = premium.quotation_insurance_vehicle.vehicle
        return issuance, ivr, vehicle

    def get_initial(self):
        initial = super().get_initial()
        plan = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self._get_data()[0].insurance_plan.id
        )
        # conversion a porcentaje
        initial["plan_commission_percentage"] = round(plan.commission * 100, 1)
        initial["final_validity"] = self._get_data()[0].final_validity
        return initial

    def form_valid(self, form):
        form.instance.vehicle_id = self._get_data()[2].id
        # obtener porcentaje actual de comisión de la aseguradora
        form.instance.tax_percentage = self._get_data()[1].tax
        form.instance.emission_right_percentage = self._get_data()[
            1
        ].emission_right
        raw_percentage = form.cleaned_data["plan_commission_percentage"]
        form.instance.plan_commission_percentage = raw_percentage / 100
        # obtener comisión actual del vendedor
        form.instance.seller_commission_percentage = self._get_data()[
            0
        ].consultant_seller.commission_rate.new_sale
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Crear endoso con movimiento de prima"
        context["body"] = "Formulario endoso vehicular"
        return context


class UpdateEndorsementSupportView(
    rrgg_mixins.RrggBootstrapDisplayMixin, LoginRequiredMixin, UpdateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/endorsement_form.html"
    model = rrgg.models.EndorsementVehicle
    fields = [
        "insured_amount",
        "net_premium",
        "rate",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "currency",
        "payment_method",
        "detail",
    ]
    pk_url_kwarg = "endorsement_id"

    def get_initial(self):
        initial = super().get_initial()
        # conversion a porcentaje
        initial["plan_commission_percentage"] = round(
            self.object.plan_commission_percentage * 100, 1
        )
        initial["final_validity"] = self.object.final_validity
        return initial

    def form_valid(self, form):
        # obtener porcentaje actual de comisión de la aseguradora
        form.instance.tax_percentage = self.object.tax_percentage
        form.instance.emission_right_percentage = (
            self.object.emission_right_percentage
        )
        raw_percentage = form.cleaned_data["plan_commission_percentage"]
        form.instance.plan_commission_percentage = raw_percentage / 100
        # obtener comisión actual del vendedor
        form.instance.seller_commission_percentage = (
            self.object.seller_commission_percentage
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Actualizar endoso"
        context["body"] = "Formulario endoso vehicular"
        return context


class EndorsementDetailSupportView(LoginRequiredMixin, DetailView):
    template_name = (
        "rrggweb/issuance/insurance/vehicle/endorsement_detail.html"
    )
    model = rrgg.models.EndorsementVehicle
    fields = "__all__"

    def _get_data(self):
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        return premium.quotation_insurance_vehicle.vehicle

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["subtitle"] = "Detalle de endoso"
        context["kcs_commission_percentage"] = round(
            self.object.plan_commission_percentage * 100, 1
        )
        context["issuance"] = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        return context


# HOME


class HomeView(TemplateView):
    template_name = "rrggweb/home.html"


# ------------------------------

# QUOTATION MAIN


class QuotationView(TemplateView):
    template_name = "rrggweb/quotation.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["seguros"] = [
            SeguroItem(
                "Seguro de vehicular",
                urls.reverse(
                    "rrggweb:quotation:insurance:vehicle:list",
                    kwargs={"registrar_id": self.kwargs["registrar_id"]},
                ),
            ),
            SeguroItem("Seguro de vida", ""),
            SeguroItem("Seguro de accidentes", ""),
            SeguroItem("Seguro de salud", ""),
        ]
        return context


class QIVListView(QuotationListSupportView):
    template_name = "rrggweb/quotation/insurance/vehicle/list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["new_register"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["previous_list"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class QIVDetailView(DetailQuotationSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["subtitle"] = "Detalle de la cotización"
        context["manage_premiums"] = True
        context["seller"] = self.object.consultant_seller
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["update_owner"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "owner_id": context["owner"].owner.id,
                    "quotation_id": self.object.id,
                },
            )
        else:
            pass
        if isinstance(context["customer"].pick, rrgg.models.NaturalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_natural_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "natural_person_id": context["customer"].pick.id,
                    "quotation_id": self.object.id,
                },
            )
        elif isinstance(context["customer"].pick, rrgg.models.LegalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_legal_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "legal_person_id": context["customer"].pick.id,
                    "quotation_id": self.object.id,
                },
            )
        else:
            pass
        context["update"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:update",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["update_vehicle"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:update_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "vehicle_id": self.object.vehicle.id,
                "quotation_id": self.object.id,
            },
        )
        context["update_seller"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:update_seller",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["create_premium"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create_premium",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["report_xlsx"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:report_xlsx",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["report_pdf"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:report_pdf",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class QIVUpdateSellerView(UpdateSellerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        return context


class QIVUpdateNaturalPersonView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class QIVUpdateLegalPersonView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class QIVUpdateVehicleView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["pretty_style"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class QIVUpdateOwnerView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class QIVUpdateView(UpdateQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        return context


class QIVCreatePremiumView(CreatePremiumSupportView):
    def get_success_url(self):
        return urls.reverse_lazy(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=self.kwargs["quotation_id"],
        )
        context["insured_amount"] = quotation.insured_amount
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class QIVUpdatePremiumView(UpdatePremiumSupportView):
    def get_success_url(self):
        return urls.reverse_lazy(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
            },
        )
        return context


class QIVDeletePremiumView(DeletePremiumSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
            },
        )
        return context


# REPORTS


class QIVReportXlsxView(View):
    def get(self, request, *args, **kwargs):
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=kwargs["quotation_id"],
        )
        workbook = self.create_workbook(quotation)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats"
            + "-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename=report_quotations.xlsx"
        )
        workbook.save(response)
        return response

    def create_workbook(self, quotation):
        from openpyxl import load_workbook

        wb = load_workbook(filename="rrggweb/resources/report_quotations.xlsx")
        ws = wb.active
        ws["C11"] = quotation.insured_amount
        ws["I6"] = quotation.created.strftime("%d/%m/%Y")
        vehicle = quotation.vehicle
        ws["C7"] = vehicle.brand
        ws["C8"] = vehicle.vehicle_model
        ws["C9"] = vehicle.fabrication_year
        ws["C10"] = vehicle.use_type.name
        ws["C6"] = str(quotation.customer)

        for premium in quotation.premiums.all():
            ratio = premium.insurance_vehicle_ratio
            insurance_vehicle = ratio.insurance_vehicle
            if premium.amount > 0 and insurance_vehicle.id == 1:
                ws["E13"] = premium.amount
                ws["E14"] = ws["E13"].value * ratio.emission_right
                ws["E15"] = (ws["E13"].value + ws["E14"].value) * ratio.tax
                ws["E16"] = ws["E13"].value + ws["E14"].value + ws["E15"].value
            if premium.amount > 0 and insurance_vehicle.id == 2:
                ws["G13"] = premium.amount
                ws["G14"] = ws["G13"].value * ratio.emission_right
                ws["G15"] = (ws["G13"].value + ws["G14"].value) * ratio.tax
                ws["G16"] = ws["G13"].value + ws["G14"].value + ws["G15"].value
            if premium.amount > 0 and insurance_vehicle.id == 3:
                ws["I13"] = premium.amount
                ws["I14"] = ws["I13"].value * ratio.emission_right
                ws["I15"] = (ws["I13"].value + ws["I14"].value) * ratio.tax
                ws["I16"] = ws["I13"].value + ws["I14"].value + ws["I15"].value
            if premium.amount > 0 and insurance_vehicle.id == 4:
                ws["K13"] = premium.amount
                ws["K14"] = ws["K13"].value * ratio.emission_right
                ws["K15"] = (ws["K13"].value + ws["K14"].value) * ratio.tax
                ws["K16"] = ws["K13"].value + ws["K14"].value + ws["K15"].value
            if premium.amount > 0 and insurance_vehicle.id == 5:
                ws["M13"] = premium.amount
                ws["M14"] = ws["M13"].value * ratio.emission_right
                ws["M15"] = (ws["M13"].value + ws["M14"].value) * ratio.tax
                ws["M16"] = ws["M13"].value + ws["M14"].value + ws["M15"].value

        return wb


class QIVReportPdfView(View):
    def get(self, request, *args, **kwargs):
        from django.template.loader import render_to_string
        from weasyprint import HTML

        templatename = "rrggweb/quotation/insurance/vehicle/report.html"
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=kwargs["quotation_id"],
        )
        premiums = quotation.premiums.all().order_by(
            "insurance_vehicle_ratio__insurance_vehicle__id"
        )
        my_range = range(1, 6)

        html_string = render_to_string(
            templatename,
            {
                "quotation": quotation,
                "premiums": premiums,
                "my_range": my_range,
            },
        )

        pdf_file = HTML(string=html_string).write_pdf()

        response = HttpResponse(pdf_file, content_type="application/pdf")
        response["Content-Disposition"] = (
            "attachment; filename=report_quotations.pdf"
        )

        return response


# QUOTATION - STEP


class QIVSelectRoleFormView(SelectRoleSupportView):
    def form_valid(self, form: forms.RoleForm):
        role = form.cleaned_data["roles"]
        self.success_url = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_seller",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": role.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 1
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class QIVSelectSellerFormView(SelectSellerSupportView):
    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 1
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class QIVSearchCustomerView(SearchCustomerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        natural_customer_exists = (
            rrgg.models.CustomerMembership.objects.filter(
                natural_person__document_number=document_number
            ).exists()
        )
        legal_customer_exists = rrgg.models.CustomerMembership.objects.filter(
            legal_person__document_number=document_number
        ).exists()
        if natural_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                natural_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:search_vehicle",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        elif legal_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                legal_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:search_vehicle",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        else:
            select_customer_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:select_customer",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = (
                f"{select_customer_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class QIVSelectCustomerFormView(SelectCustomerSupportView):
    def form_valid(self, form: forms.SelectCustomerForm):
        type_customer = form.cleaned_data["type_customer"]
        dn = self.request.GET.get("document_number", "")
        if type_customer == "persona_natural":
            create_customer_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create_natural_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        elif type_customer == "persona_jurídica":
            create_customer_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create_legal_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        else:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["subtitle"] = "Definir contratante"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class QIVCreateNaturalPersonView(CreateNaturalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            natural_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class QIVUpdateNaturalPersonStepView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class QIVCreateLegalPersonView(CreateLegalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            legal_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )

        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:select_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class QIVUpdateLegalPersonStepView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 2
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class QIVSearchVehicleView(SearchVehicleSupportView):
    def form_valid(self, form):
        plate = form.cleaned_data["plate"]
        vehicle_exists = rrgg.models.Vehicle.objects.filter(
            plate=plate
        ).exists()
        # si existe el vehículo, se debe verificar si tiene propietario
        if vehicle_exists:
            vehicle = rrgg.models.Vehicle.objects.get(plate=plate)
            vehicle_ownership_exists = (
                rrgg.models.VehicleOwnership.objects.filter(
                    vehicle__plate=plate
                ).exists()
            )
            if vehicle_ownership_exists:
                vehicle_ownership = rrgg.models.VehicleOwnership.objects.get(
                    vehicle__plate=plate
                )
                # si tiene propietario se debe verificar si
                # es de tipo contratante
                if isinstance(
                    vehicle_ownership.pick, rrgg.models.CustomerMembership
                ):
                    customer = shortcuts.get_object_or_404(
                        rrgg.models.CustomerMembership,
                        id=self.kwargs["customer_id"],
                    )
                    # si es de tipo contratante, se debe verificar
                    # si es el mismo
                    if vehicle_ownership.pick == customer:
                        self.success_url = urls.reverse(
                            "rrggweb:quotation:insurance:vehicle:create",
                            kwargs={
                                "registrar_id": self.kwargs["registrar_id"],
                                "seller_id": self.kwargs["seller_id"],
                                "customer_id": self.kwargs["customer_id"],
                                "vehicle_id": vehicle.id,
                            },
                        )
                    else:
                        form.add_error(
                            "plate",
                            "El contratante no es dueño del vehículo",
                        )
                        return super().form_invalid(form)
                else:
                    # TODO: falta evaluar si el propietario registrado
                    # es el actual
                    self.success_url = urls.reverse(
                        "rrggweb:quotation:insurance:vehicle:create",
                        kwargs={
                            "registrar_id": self.kwargs["registrar_id"],
                            "seller_id": self.kwargs["seller_id"],
                            "customer_id": self.kwargs["customer_id"],
                            "vehicle_id": vehicle.id,
                        },
                    )
            else:
                self.success_url = urls.reverse(
                    "rrggweb:quotation:insurance:vehicle:define_owner",
                    kwargs={
                        "registrar_id": self.kwargs["registrar_id"],
                        "seller_id": self.kwargs["seller_id"],
                        "customer_id": self.kwargs["customer_id"],
                        "vehicle_id": vehicle.id,
                    },
                )
        else:
            create_vehicle_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create_vehicle",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                },
            )
            self.success_url = f"{create_vehicle_url}?plate={plate}"
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 3
        context["final_step"] = 6
        if isinstance(context["customer"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                (
                    "rrggweb:quotation:insurance:vehicle:"
                    "update_natural_person_step"
                ),
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "natural_person_id": context["customer"].pick.id,
                },
            )
        elif isinstance(context["customer"].pick, rrgg.models.LegalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_legal_person_step",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "legal_person_id": context["customer"].pick.id,
                },
            )
        else:
            pass
        return context


class QIVCreateVehicleView(CreateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:define_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 3
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class QIVUpdateVehicleStepView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:define_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 3
        context["final_step"] = 6
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class QIVDefineOwnerView(DefineOwnerSupportView):
    def form_valid(self, form):
        # eliminar la relación de propiedad si existe
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        vehicle_ownership_exists = rrgg.models.VehicleOwnership.objects.filter(
            vehicle=vehicle
        ).exists()
        if vehicle_ownership_exists:
            vehicle.ownership.delete()
        else:
            pass
        is_customer_owner = form.cleaned_data["is_owner"]
        if is_customer_owner:
            customer = shortcuts.get_object_or_404(
                rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                customer=customer, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            self.success_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:search_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 4
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:update_vehicle_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class QIVSearchOwnerView(SearchOwnerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        if not re.match("^[0-9]+$", document_number):
            form.add_error(
                "document_number",
                "El número de documento debe contener solo números.",
            )
            return super().form_invalid(form)
        owner_exists = rrgg.models.NaturalPerson.objects.filter(
            document_number=document_number
        ).exists()
        if owner_exists:
            # si existe la persona en db, se crea la relación
            owner = rrgg.models.NaturalPerson.objects.get(
                document_number=document_number
            )
            vehicle = shortcuts.get_object_or_404(
                rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                owner=owner, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            create_owner_url = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:create_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
            self.success_url = (
                f"{create_owner_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 4
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:define_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class QIVCreateOwnerView(CreateOwnerSupportView):
    def get_success_url(self):
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        rrgg.models.VehicleOwnership.objects.create(
            owner=self.object, vehicle=vehicle
        )
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 4
        context["final_step"] = 6
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:search_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class QIVUpdateOwnerStepView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["initial_step"] = 4
        context["final_step"] = 6
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:define_owner",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class QIVCreateView(CreateQuotationSupportView):
    def form_valid(self, form):
        form.instance.source = "quotation"
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create_premiums",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["subtitle"] = "Registrar cotización"
        context["initial_step"] = 5
        context["final_step"] = 6
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_owner_step",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                    "owner_id": context["owner"].owner.id,
                },
            )
        else:
            context["previous_page"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:define_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return context


class QIVUpdateStepView(UpdateQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:quotation:insurance:vehicle:create_premiums",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["initial_step"] = 5
        context["final_step"] = 6
        context["seller"] = self.object.consultant_seller
        context["customer"] = self.object.customer
        context["vehicle"] = self.object.vehicle
        context["owner"] = context["vehicle"].ownership
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:update_owner_step",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.object.consultant_seller.id,
                    "customer_id": self.object.customer.id,
                    "vehicle_id": self.object.vehicle.id,
                    "owner_id": context["owner"].owner.id,
                },
            )
        else:
            context["previous_page"] = urls.reverse(
                "rrggweb:quotation:insurance:vehicle:define_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.object.consultant_seller.id,
                    "customer_id": self.object.customer.id,
                    "vehicle_id": self.object.vehicle.id,
                },
            )
        return context


class QIVPremiumsFormView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/create_premiums.html"

    def form_valid(self, form):
        form.save()
        return super().form_valid(form)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["queryset"] = rrgg.models.InsuranceVehicle.objects.none()
        return kwargs

    def get_form(self):
        form_class = modelformset_factory(
            rrgg.models.QuotationInsuranceVehiclePremium,
            extra=rrgg.models.InsuranceVehicle.objects.count(),
            fields=[
                "insurance_vehicle_ratio",
                "quotation_insurance_vehicle",
                "amount",
                "rate",
            ],
        )
        formset = super().get_form(form_class)
        # Mantener el orden: Ver ABC123
        insurance_vehicles = rrgg.models.InsuranceVehicle.objects.order_by(
            "name"
        )
        for form, insurance_vehicle in zip(formset, insurance_vehicles):
            insurance_vehicle_ratio = form.fields["insurance_vehicle_ratio"]
            quotation_insurance_vehicle = form.fields[
                "quotation_insurance_vehicle"
            ]
            insurance_vehicle_ratio.initial = insurance_vehicle.last_ratio
            quotation_insurance_vehicle.initial = shortcuts.get_object_or_404(
                rrgg.models.QuotationInsuranceVehicle,
                id=self.kwargs["quotation_id"],
            )
            form.tax_percentage = insurance_vehicle_ratio.initial.tax
            form.emission_right_percentage = (
                insurance_vehicle_ratio.initial.emission_right
            )
            insurance_vehicle_ratio.widget = forms.forms.HiddenInput()
            quotation_insurance_vehicle.widget = forms.forms.HiddenInput()

        return formset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["step"] = True
        context["title"] = "COTIZACIÓN VEHICULAR"
        context["subtitle"] = "Registrar primas de aseguradoras"
        context["quotation"] = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=self.kwargs["quotation_id"],
        )
        context["insurances"] = rrgg.models.InsuranceVehicle.objects.all()
        context["seller"] = context["quotation"].consultant_seller
        context["customer"] = context["quotation"].customer
        context["vehicle"] = context["quotation"].vehicle
        context["owner"] = context["vehicle"].ownership
        context["previous_page"] = urls.reverse(
            "rrggweb:quotation:insurance:vehicle:update_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        # Mantener el orden: Ver ABC123
        insurance_vehicles = rrgg.models.InsuranceVehicle.objects.order_by(
            "name"
        )
        last_ratios = (
            insurance_vehicle.last_ratio
            for insurance_vehicle in insurance_vehicles
        )
        context["last_ratio_forms"] = zip(last_ratios, context["form"])

        return context

    def get_success_url(self):
        return urls.reverse_lazy(
            "rrggweb:quotation:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )


# ------------------------------

# ISSUANCE - MAIN


class IssuanceView(TemplateView):
    template_name = "rrggweb/issuance.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["seguros"] = [
            SeguroItem(
                "Seguro de vehicular",
                urls.reverse(
                    "rrggweb:issuance:insurance:vehicle:list",
                    kwargs={"registrar_id": self.kwargs["registrar_id"]},
                ),
            ),
            SeguroItem("Seguro de vida", ""),
            SeguroItem("Seguro de accidentes", ""),
            SeguroItem("Seguro de salud", ""),
        ]
        return context


class IIVListView(IssuanceListSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["new_register"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_record_type",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["type"] = "initial"
        return context


class IIVUpdateIssuanceDetailView(UpdateIssuanceSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["step"] = False
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVDetailView(LoginRequiredMixin, DetailView):
    template_name = "rrggweb/issuance/insurance/vehicle/detail.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    pk_url_kwarg = "issuance_id"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["subtitle"] = "Detalle de la emisión"
        context["premiums"] = self.object.quotation_vehicle_premiums.all()
        context["seller"] = self.object.consultant_seller
        context["kcs_commission_percentage"] = round(
            self.object.plan_commission_percentage * 100, 1
        )
        context["create_document"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ed",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )
        context["documents"] = self.object.documents.all()
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["anular"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_status",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVUpdateStatusDetailView(FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.IssuanceStatusForm

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["comment"].required = False
        return form

    def form_valid(self, form: forms.IssuanceStatusForm):
        issuance = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        data = form.cleaned_data
        issuance.status = data["status"]
        issuance.comment = data["comment"]
        issuance.save()
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["subtitle"] = "Gestionar estado de emisión"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVAddDocumentDetailView(AddDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ed",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDeleteDocumentDetailView(DeleteDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["return"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVCreateEndorsementDetailView(CreateEndorsementSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:endorsement_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDetailEndorsementDetailView(EndorsementDetailSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVUpdateEndorsementDetailView(UpdateEndorsementSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:endorsement_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDetailVehicleDetailView(LoginRequiredMixin, DetailView):
    template_name = "rrggweb/issuance/insurance/vehicle/detail_vehicle.html"
    model = rrgg.models.Vehicle
    fields = "__all__"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - BIEN ASEGURADO"
        context["subtitle"] = "Detalle del bien asegurado"
        context["issuance"] = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        context["premium"] = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVCreateEndorsementVehicleView(CreateEndorsementSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:endorsement_detail_v",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - BIEN ASEGURADO"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:vehicle_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self._get_data()[2].id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVUpdateEndorsementVehicleView(UpdateEndorsementSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:endorsement_detail_v",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - BIEN ASEGURADO"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:vehicle_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.vehicle.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDetailEndorsementVehicleView(EndorsementDetailSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - BIEN ASEGURADO"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:vehicle_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "pk": self.object.vehicle.id,
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


# ISSUANCE - STEP


class IIVDefineRegistrationTypeView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.SelectVehicleRegistrationTypeForm

    def form_valid(self, form):
        vehicle_registration_type = form.cleaned_data[
            "vehicle_registration_type"
        ]
        if vehicle_registration_type == "new_sale":
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:define_new_sale",
                kwargs={"registrar_id": self.kwargs["registrar_id"]},
            )
        elif vehicle_registration_type == "renewal":
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:list_renewal",
                kwargs={"registrar_id": self.kwargs["registrar_id"]},
            )
        else:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["subtitle"] = "Definir tipo de registro"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVDefineNewSaleView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.DefineNewSaleForm

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["has_quote"].required = False
        return form

    def form_valid(self, form):
        has_quote = form.cleaned_data["has_quote"]
        if has_quote:
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:list_quotations",
                kwargs={"registrar_id": self.kwargs["registrar_id"]},
            )
        else:
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:select_role_ns",
                kwargs={"registrar_id": self.kwargs["registrar_id"]},
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR"
        context["subtitle"] = "Definir venta nueva"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_record_type",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


# ISSUANCE - QUOTATION - MAIN


class IIVListQuotationView(QuotationListSupportView):
    template_name = "rrggweb/issuance/insurance/vehicle/list_quotations.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_list"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_quotations",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_new_sale",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVQuotationDetailView(DetailQuotationSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["subtitle"] = "Detalle del registro vehicular"
        context["manage_premiums"] = False
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["update_owner"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "owner_id": context["owner"].owner.id,
                    "quotation_id": self.object.id,
                },
            )
        else:
            pass
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_quotations",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        if isinstance(context["customer"].pick, rrgg.models.NaturalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_natural_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "natural_person_id": context["customer"].pick.id,
                    "quotation_id": self.object.id,
                },
            )
        elif isinstance(context["customer"].pick, rrgg.models.LegalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_legal_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "legal_person_id": context["customer"].pick.id,
                    "quotation_id": self.object.id,
                },
            )
        else:
            pass

        context["update"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_quotation",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        context["update_vehicle"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "vehicle_id": self.object.vehicle.id,
                "quotation_id": self.object.id,
            },
        )
        return context


class IIVUpdateNaturalPersonQuotationView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class IIVUpdateLegalPersonQuotationView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class IIVUpdateVehicleQuotationView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class IIVUpdateOwnerQuotationView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class IIVUpdateQuotationView(UpdateQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
            },
        )
        return context


# ISSUANCE - QUOTATION - STEP


class IIVSelectRoleQuotationFormView(SelectRoleSupportView):
    def form_valid(self, form: forms.RoleForm):
        role = form.cleaned_data["roles"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_seller_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "role_id": role.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 1
        context["final_step"] = 4
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        quotation = premium.quotation_insurance_vehicle
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:quotation_detail",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": quotation.id,
            },
        )
        return context


class IIVSelectSellerQuotationView(SelectSellerSupportView):
    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 1
        context["final_step"] = 4
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_role_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
            },
        )
        return context


class IIVSelectPlanQuotationView(SelectPlanSupportView):
    def form_valid(self, form: forms.InsurancePlanForm):
        plan = form.cleaned_data["planes de seguro"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
                "plan_id": plan.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_role_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
            },
        )
        return context


class IIVCreateStepQuotationView(CreateIssuanceSupportView):
    def form_valid(self, form):
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        quotation = premium.quotation_insurance_vehicle
        if quotation.expired:
            messages.warning(self.request, "Esta cotización ya expiró")
            return self.form_invalid(form)
        else:
            quotation.source = "new_sale-quotation"
            quotation.save()
            issuance = form.save(commit=False)
            issuance.issuance_type_id = 1
            issuance.consultant_registrar_id = self.kwargs["registrar_id"]
            issuance.consultant_seller_id = self.kwargs["seller_id"]
            issuance.insurance_plan_id = self.kwargs["plan_id"]

            raw_percentage = form.cleaned_data["plan_commission_percentage"]
            issuance.plan_commission_percentage = raw_percentage / 100
            seller = shortcuts.get_object_or_404(
                rrgg.models.Consultant, id=self.kwargs["seller_id"]
            )
            # TODO: comisión del vendedor por ahora solo tomara de referencia
            # la comisión actual registrada en el momento de la creación
            issuance.seller_commission_percentage = (
                seller.commission_rate.new_sale
            )
            issuance.save()
            issuance.quotation_vehicle_premiums.add(premium)
            return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 3
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateStepQuotationView(UpdateIssuanceSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 3
        context["final_step"] = 4
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.object.quotation_vehicle_premium.id,
                "seller_id": self.object.consultant_seller.id,
            },
        )
        return context


class IIVAddDocumentQuotationStepView(AddDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 4
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_step_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDeleteDocumentQuotationStepView(DeleteDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - COTIZACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 4
        context["return"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_q",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


# ISSUANCE - NEW SALE - STEP


class IIVSelectRoleNewSaleView(SelectRoleSupportView):
    def form_valid(self, form: forms.RoleForm):
        role = form.cleaned_data["roles"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_seller_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": role.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 1
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_new_sale",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVSelectSellerNewSaleView(SelectSellerSupportView):
    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 1
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_role_ns",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVSearchCustomerNewSaleView(SearchCustomerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        natural_customer_exists = (
            rrgg.models.CustomerMembership.objects.filter(
                natural_person__document_number=document_number
            ).exists()
        )
        legal_customer_exists = rrgg.models.CustomerMembership.objects.filter(
            legal_person__document_number=document_number
        ).exists()
        if natural_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                natural_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        elif legal_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                legal_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        else:
            select_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:select_customer_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = (
                f"{select_customer_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_seller_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": context["seller"].role.id,
            },
        )
        return context


class IIVSelectCustomerNewSaleView(SelectCustomerSupportView):
    def form_valid(self, form: forms.SelectCustomerForm):
        type_customer = form.cleaned_data["type_customer"]
        dn = self.request.GET.get("document_number", "")
        if type_customer == "persona_natural":
            create_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_natural_person_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        elif type_customer == "persona_jurídica":
            create_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_legal_person_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        else:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateNaturalPersonNewSaleView(CreateNaturalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            natural_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateLegalPersonNewSaleView(CreateLegalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            legal_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateNaturalPersonNewSaleStepView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateLegalPersonNewSaleStepView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 3
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVPremiumListNewSaleStepView(PremiumListSupportView):
    template_name = "rrggweb/issuance/insurance/vehicle/list_premium.html"

    def get_queryset(self):
        customer_id = self.kwargs["customer_id"]
        seller_id = self.kwargs["seller_id"]
        return rrgg.models.QuotationInsuranceVehiclePremium.objects.filter(
            quotation_insurance_vehicle__customer_id=customer_id,
            quotation_insurance_vehicle__consultant_seller_id=seller_id,
            in_progress=True,
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 3
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["new_register"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        context["select_currency_ivr"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_currency_ivr_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        if isinstance(context["customer"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_natural_person_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "natural_person_id": context["customer"].pick.id,
                },
            )
        elif isinstance(context["customer"].pick, rrgg.models.LegalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_legal_person_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "legal_person_id": context["customer"].pick.id,
                },
            )
        else:
            pass
        return context


class IIVUpdatePremiumNewSaleStepView(UpdatePremiumQuotationSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": (
                    self.object.quotation_insurance_vehicle.consultant_seller.id  # noqa: E501
                ),
                "customer_id": (
                    self.object.quotation_insurance_vehicle.customer.id
                ),
            },
        )
        context["initial_step"] = 3
        context["final_step"] = 9
        return context


class IIVDeletePremiumNewSaleStepView(DeletePremiumQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        context["initial_step"] = 3
        context["final_step"] = 9
        return context


class IIVSearchVehicleNewSaleView(SearchVehicleSupportView):
    def form_valid(self, form):
        plate = form.cleaned_data["plate"]
        vehicle_exists = rrgg.models.Vehicle.objects.filter(
            plate=plate
        ).exists()
        if vehicle_exists:
            vehicle = rrgg.models.Vehicle.objects.get(plate=plate)
            vehicle_ownership_exists = (
                rrgg.models.VehicleOwnership.objects.filter(
                    vehicle__plate=plate
                ).exists()
            )
            if vehicle_ownership_exists:
                if isinstance(
                    vehicle.ownership.pick, rrgg.models.CustomerMembership
                ):
                    customer = shortcuts.get_object_or_404(
                        rrgg.models.CustomerMembership,
                        id=self.kwargs["customer_id"],
                    )
                    if vehicle.ownership.pick == customer:
                        self.success_url = urls.reverse(
                            (
                                "rrggweb:issuance:insurance:"
                                "vehicle:create_premium_ns"
                            ),
                            kwargs={
                                "registrar_id": self.kwargs["registrar_id"],
                                "seller_id": self.kwargs["seller_id"],
                                "customer_id": self.kwargs["customer_id"],
                                "vehicle_id": vehicle.id,
                            },
                        )
                    else:
                        form.add_error(
                            "plate",
                            "El contratante no es dueño del vehículo",
                        )
                        return super().form_invalid(form)
                else:
                    # toca evaluar si el propietario registrado es el actual
                    self.success_url = urls.reverse(
                        (
                            "rrggweb:issuance:insurance:"
                            "vehicle:create_quotation_ns"
                        ),
                        kwargs={
                            "registrar_id": self.kwargs["registrar_id"],
                            "seller_id": self.kwargs["seller_id"],
                            "customer_id": self.kwargs["customer_id"],
                            "vehicle_id": vehicle.id,
                        },
                    )
            else:
                self.success_url = urls.reverse(
                    "rrggweb:issuance:insurance:vehicle:define_owner_ns",
                    kwargs={
                        "registrar_id": self.kwargs["registrar_id"],
                        "seller_id": self.kwargs["seller_id"],
                        "customer_id": self.kwargs["customer_id"],
                        "vehicle_id": vehicle.id,
                    },
                )
        else:
            create_vehicle_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_vehicle_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                },
            )
            self.success_url = f"{create_vehicle_url}?plate={plate}"
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVCreateVehicleNewSaleView(CreateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVUpdateVehicleNewSaleStepView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVDefineOwnerNewSaleView(DefineOwnerSupportView):
    def form_valid(self, form):
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        vehicle_ownership_exists = rrgg.models.VehicleOwnership.objects.filter(
            vehicle=vehicle
        ).exists()
        if vehicle_ownership_exists:
            vehicle.ownership.delete()
        else:
            pass

        is_customer_owner = form.cleaned_data["is_owner"]
        if is_customer_owner:
            customer = shortcuts.get_object_or_404(
                rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                customer=customer, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_premium_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:search_owner_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_vehicle_step_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVSearchOwnerNewSaleView(SearchOwnerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        if not re.match("^[0-9]+$", document_number):
            form.add_error(
                "document_number",
                "El número de documento debe contener solo números.",
            )
            return super().form_invalid(form)
        owner_exists = rrgg.models.NaturalPerson.objects.filter(
            document_number=document_number
        ).exists()
        if owner_exists:
            owner = rrgg.models.NaturalPerson.objects.get(
                document_number=document_number
            )
            vehicle = shortcuts.get_object_or_404(
                rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                owner=owner, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_premium_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            create_owner_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_owner_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
            self.success_url = (
                f"{create_owner_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVCreateOwnerNewSaleView(CreateOwnerSupportView):
    def get_success_url(self):
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        rrgg.models.VehicleOwnership.objects.create(
            owner=self.object, vehicle=vehicle
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_owner_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVUpdateOwnerNewSaleStepView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVCreatePremiumNewSaleView(CreatePremiumQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 6
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        context["owner"] = context["vehicle"].ownership
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_owner_step_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                    "owner_id": context["owner"].owner.id,
                },
            )
        else:
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:define_owner_ns",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return context


class IIVSelectCurrencyInsuranceNewSaleView(
    SelectCurrencyInsuranceSupportView
):
    def get_queryset(self):
        customer_id = self.kwargs["customer_id"]
        seller_id = self.kwargs["seller_id"]
        return rrgg.models.QuotationInsuranceVehiclePremium.objects.filter(
            quotation_insurance_vehicle__customer_id=customer_id,
            quotation_insurance_vehicle__consultant_seller_id=seller_id,
            in_progress=True,
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        qivp = self.get_queryset().first()
        kwargs["currency_id"] = qivp.quotation_insurance_vehicle.currency.id
        kwargs["ivr_id"] = qivp.insurance_vehicle_ratio.id
        return kwargs

    def form_valid(self, form: forms.CurrencyInsuranceForm):
        currency = form.cleaned_data["moneda"]
        ivr = form.cleaned_data["aseguradoras"]

        for premium in self.get_queryset():
            quotation = premium.quotation_insurance_vehicle
            quotation.currency = currency
            quotation.save()
            premium.insurance_vehicle_ratio = ivr
            premium.emission_right_percentage = ivr.emission_right
            premium.tax_percentage = ivr.tax
            premium.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA MÚLTIPLE"
        context["initial_step"] = 6
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["risk_selector"] = forms.RiskForm(
            risk_id=self.get_queryset()
            .first()
            .quotation_insurance_vehicle.risk_id
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_premiums_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVSelectPlanNewSaleView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/issuance/insurance/vehicle/basic_form.html"
    form_class = forms.InsurancePlanForm

    def _get_data(self):
        customer_id = self.kwargs["customer_id"]
        seller_id = self.kwargs["seller_id"]
        return rrgg.models.QuotationInsuranceVehiclePremium.objects.filter(
            quotation_insurance_vehicle__customer_id=customer_id,
            quotation_insurance_vehicle__consultant_seller_id=seller_id,
            in_progress=True,
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        qivp = self._get_data().first()
        ivr = qivp.insurance_vehicle_ratio
        riv = shortcuts.get_object_or_404(
            rrgg.models.RiskInsuranceVehicle,
            insurance_vehicle_id=ivr.insurance_vehicle_id,
            risk_id=qivp.quotation_insurance_vehicle.risk_id,
        )
        kwargs["riv_id"] = riv.id
        return kwargs

    def form_valid(self, form: forms.InsurancePlanForm):
        plan = form.cleaned_data["planes de seguro"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "plan_id": plan.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA MULTIPLE"
        context["subtitle"] = "Seleccionar plan de seguro"
        context["initial_step"] = 7
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        riv = shortcuts.get_object_or_404(
            rrgg.models.RiskInsuranceVehicle,
            id=self.get_form_kwargs()["riv_id"],
        )
        context["risk_selector"] = forms.RiskForm(risk_id=riv.risk_id)
        context["insurance_vehicle_selector"] = forms.InsuranceVehicleForm(
            iv_id=riv.insurance_vehicle_id
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_currency_ivr_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVCreateStepNewSaleView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/form_multiple.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    fields = [
        "policy",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "payment_method",
    ]

    def get_initial(self):
        initial = super().get_initial()
        plan = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self.kwargs["plan_id"]
        )
        kcs_commission_percentage = round(plan.commission * 100, 1)
        initial["plan_commission_percentage"] = kcs_commission_percentage
        return initial

    def _get_data(self):
        customer_id = self.kwargs["customer_id"]
        seller_id = self.kwargs["seller_id"]
        return rrgg.models.QuotationInsuranceVehiclePremium.objects.filter(
            quotation_insurance_vehicle__customer_id=customer_id,
            quotation_insurance_vehicle__consultant_seller_id=seller_id,
            in_progress=True,
        )

    def form_valid(self, form):
        premiums = self._get_data()
        if premiums.first().quotation_insurance_vehicle.expired:
            messages.warning(self.request, "Esta cotización ya expiró")
            return self.form_invalid(form)
        else:
            issuance = form.save(commit=False)
            issuance.issuance_type_id = 1
            issuance.consultant_registrar_id = self.kwargs["registrar_id"]
            issuance.consultant_seller_id = self.kwargs["seller_id"]
            issuance.insurance_plan_id = self.kwargs["plan_id"]

            raw_percentage = form.cleaned_data["plan_commission_percentage"]
            issuance.plan_commission_percentage = raw_percentage / 100
            seller = shortcuts.get_object_or_404(
                rrgg.models.Consultant, id=self.kwargs["seller_id"]
            )
            # TODO: comisión del vendedor por ahora solo tomara de referencia
            # la comisión actual registrada en el momento de la creación
            issuance.seller_commission_percentage = (
                seller.commission_rate.new_sale
            )
            issuance.save()
            for premium in premiums:
                premium.in_progress = False
                premium.save()
                issuance.quotation_vehicle_premiums.add(premium)
            return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["subtitle"] = "Crear emisión"
        context["initial_step"] = 8
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["premiums"] = self._get_data()
        context["insured_amount"] = sum(
            p.quotation_insurance_vehicle.insured_amount
            for p in context["premiums"]
        )
        context["net_premium"] = sum(p.amount for p in context["premiums"])
        context["rate"] = round(
            sum(p.rate for p in context["premiums"])
            / len(context["premiums"]),
            2,
        )
        context["emission_right"] = round(
            context["net_premium"]
            * context["premiums"].first().emission_right_percentage,
            2,
        )
        context["commercial_premium"] = (
            context["net_premium"] + context["emission_right"]
        )
        context["tax"] = round(
            context["commercial_premium"]
            * context["premiums"].first().tax_percentage,
            2,
        )
        context["total_premium"] = (
            context["commercial_premium"] + context["tax"]
        )
        context["insurance_plan"] = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self.kwargs["plan_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVUpdateStepNewSaleView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/form_multiple.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    fields = [
        "policy",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "payment_method",
    ]
    pk_url_kwarg = "issuance_id"

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["subtitle"] = "Editar Emisión"
        context["initial_step"] = 8
        context["final_step"] = 9
        premium = self.object.quotation_vehicle_premiums.first()
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.object.consultant_seller_id,
                "customer_id": premium.quotation_insurance_vehicle.customer.id,
            },
        )
        return context


class IIVAddDocumentNewSaleView(AddDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 9
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_step_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDeleteDocumentNewSaleView(DeleteDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - VENTA NUEVA"
        context["initial_step"] = 9
        context["final_step"] = 9
        context["return"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


# ISSUANCE - RENEWAL


class IIVRenewalListView(IssuanceListSupportView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_record_type",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["type"] = "renewal"
        context["new_renewal"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_role_nr",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


# ISSUANCE - NEW RENEWAL - STEP


class IIVSelectRoleNewRenewalView(SelectRoleSupportView):
    def form_valid(self, form: forms.RoleForm):
        role = form.cleaned_data["roles"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_seller_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": role.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 1
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_renewal",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVSelectSellerNewRenewalView(SelectSellerSupportView):
    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 1
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_role_nr",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVSearchCustomerNewRenewalView(SearchCustomerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        natural_customer_exists = (
            rrgg.models.CustomerMembership.objects.filter(
                natural_person__document_number=document_number
            ).exists()
        )
        legal_customer_exists = rrgg.models.CustomerMembership.objects.filter(
            legal_person__document_number=document_number
        ).exists()
        if natural_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                natural_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        elif legal_customer_exists:
            customer = rrgg.models.CustomerMembership.objects.get(
                legal_person__document_number=document_number
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": customer.id,
                },
            )
        else:
            select_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:select_customer_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = (
                f"{select_customer_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_seller_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": context["seller"].role.id,
            },
        )
        return context


class IIVSelectCustomerNewRenewalView(SelectCustomerSupportView):
    def form_valid(self, form: forms.SelectCustomerForm):
        type_customer = form.cleaned_data["type_customer"]
        dn = self.request.GET.get("document_number", "")
        if type_customer == "persona_natural":
            create_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_natural_person_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        elif type_customer == "persona_jurídica":
            create_customer_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_legal_person_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = f"{create_customer_url}?document_number={dn}"
        else:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateNaturalPersonNewRenewalView(CreateNaturalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            natural_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateLegalPersonNewRenewalView(CreateLegalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            legal_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateNaturalPersonNewRenewalStepView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateLegalPersonNewRenewalStepView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.object.membership.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_customer_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVSearchVehicleNewRenewalView(SearchVehicleSupportView):
    def form_valid(self, form):
        plate = form.cleaned_data["plate"]
        vehicle_exists = rrgg.models.Vehicle.objects.filter(
            plate=plate
        ).exists()
        if vehicle_exists:
            vehicle = rrgg.models.Vehicle.objects.get(plate=plate)
            vehicle_ownership_exists = (
                rrgg.models.VehicleOwnership.objects.filter(
                    vehicle__plate=plate
                ).exists()
            )
            if vehicle_ownership_exists:
                if isinstance(
                    vehicle.ownership.pick, rrgg.models.CustomerMembership
                ):
                    customer = shortcuts.get_object_or_404(
                        rrgg.models.CustomerMembership,
                        id=self.kwargs["customer_id"],
                    )
                    if vehicle.ownership.pick == customer:
                        self.success_url = urls.reverse(
                            (
                                "rrggweb:issuance:insurance:"
                                "vehicle:create_quotation_nr"
                            ),
                            kwargs={
                                "registrar_id": self.kwargs["registrar_id"],
                                "seller_id": self.kwargs["seller_id"],
                                "customer_id": self.kwargs["customer_id"],
                                "vehicle_id": vehicle.id,
                            },
                        )
                    else:
                        form.add_error(
                            "plate",
                            "El contratante no es dueño del vehículo",
                        )
                        return super().form_invalid(form)
                else:
                    # toca evaluar si el propietario registrado es el actual
                    self.success_url = urls.reverse(
                        (
                            "rrggweb:issuance:insurance:"
                            "vehicle:create_quotation_nr"
                        ),
                        kwargs={
                            "registrar_id": self.kwargs["registrar_id"],
                            "seller_id": self.kwargs["seller_id"],
                            "customer_id": self.kwargs["customer_id"],
                            "vehicle_id": vehicle.id,
                        },
                    )
            else:
                self.success_url = urls.reverse(
                    "rrggweb:issuance:insurance:vehicle:define_owner_nr",
                    kwargs={
                        "registrar_id": self.kwargs["registrar_id"],
                        "seller_id": self.kwargs["seller_id"],
                        "customer_id": self.kwargs["customer_id"],
                        "vehicle_id": vehicle.id,
                    },
                )
        else:
            create_vehicle_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_vehicle_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                },
            )
            self.success_url = f"{create_vehicle_url}?plate={plate}"
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 3
        context["final_step"] = 9
        if isinstance(context["customer"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                (
                    "rrggweb:issuance:insurance:vehicle:"
                    "update_natural_person_step_nr"
                ),
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "natural_person_id": context["customer"].pick.id,
                },
            )
        elif isinstance(context["customer"].pick, rrgg.models.LegalPerson):
            context["previous_page"] = urls.reverse(
                (
                    "rrggweb:issuance:insurance:vehicle:"
                    "update_legal_person_step_nr"
                ),
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "legal_person_id": context["customer"].pick.id,
                },
            )
        else:
            pass
        return context


class IIVCreateVehicleNewRenewalView(CreateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 3
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVUpdateVehicleNewRenewalStepView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 3
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVDefineOwnerNewRenewalView(DefineOwnerSupportView):
    def form_valid(self, form):
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        vehicle_ownership_exists = rrgg.models.VehicleOwnership.objects.filter(
            vehicle=vehicle
        ).exists()
        if vehicle_ownership_exists:
            vehicle.ownership.delete()
        else:
            pass

        is_customer_owner = form.cleaned_data["is_owner"]
        if is_customer_owner:
            customer = shortcuts.get_object_or_404(
                rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                customer=customer, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_quotation_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:search_owner_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_vehicle_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
            },
        )
        return context


class IIVSearchOwnerNewRenewalView(SearchOwnerSupportView):
    def form_valid(self, form):
        document_number = form.cleaned_data["document_number"]
        if not re.match("^[0-9]+$", document_number):
            form.add_error(
                "document_number",
                "El número de documento debe contener solo números.",
            )
            return super().form_invalid(form)
        owner_exists = rrgg.models.NaturalPerson.objects.filter(
            document_number=document_number
        ).exists()
        if owner_exists:
            owner = rrgg.models.NaturalPerson.objects.get(
                document_number=document_number
            )
            vehicle = shortcuts.get_object_or_404(
                rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
            )
            rrgg.models.VehicleOwnership.objects.create(
                owner=owner, vehicle=vehicle
            )
            self.success_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_quotation_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        else:
            create_owner_url = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:create_owner_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
            self.success_url = (
                f"{create_owner_url}?document_number={document_number}"
            )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVCreateOwnerNewRenewalView(CreateOwnerSupportView):
    def get_success_url(self):
        vehicle = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        rrgg.models.VehicleOwnership.objects.create(
            owner=self.object, vehicle=vehicle
        )
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_quotation_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_owner_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVUpdateOwnerNewRenewalStepView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_quotation_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self.kwargs["customer_id"]
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self.kwargs["vehicle_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:define_owner_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "customer_id": self.kwargs["customer_id"],
                "vehicle_id": self.kwargs["vehicle_id"],
            },
        )
        return context


class IIVCreateQuotationNewRenewalView(CreateQuotationSupportView):
    def form_valid(self, form):
        form.instance.source = "new_renewal"
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.id,
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["subtitle"] = "Registrar suma asegurada"
        context["initial_step"] = 5
        context["final_step"] = 9
        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_owner_step_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                    "owner_id": context["owner"].owner.id,
                },
            )
        else:
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:define_owner_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.kwargs["customer_id"],
                    "vehicle_id": self.kwargs["vehicle_id"],
                },
            )
        return context


class IIVUpdateQuotationStepNewRenewalView(UpdateQuotationSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 5
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["customer"] = self.object.customer
        context["vehicle"] = self.object.vehicle
        context["owner"] = context["vehicle"].ownership

        if isinstance(context["owner"].pick, rrgg.models.NaturalPerson):
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_owner_step_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.object.customer.id,
                    "vehicle_id": self.object.vehicle.id,
                    "owner_id": context["owner"].owner.id,
                },
            )
        else:
            context["previous_page"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:define_owner_nr",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                    "customer_id": self.object.customer.id,
                    "vehicle_id": self.object.vehicle.id,
                },
            )
        return context


class IIVCreatePremiumNewRenewalView(CreatePremiumSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.object.id,
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 6
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=self.kwargs["quotation_id"],
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=quotation.customer.id
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=quotation.vehicle.id
        )
        context["owner"] = quotation.vehicle.ownership
        context["risk_selector"] = forms.RiskForm(risk_id=quotation.risk_id)
        context["insured_amount"] = quotation.insured_amount
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_quotation_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.kwargs["quotation_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdatePremiumStepNewRenewalView(UpdatePremiumSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.object.id,
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 6
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        quotation = self.object.quotation_insurance_vehicle
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=quotation.customer.id
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=quotation.vehicle.id
        )
        context["owner"] = context["vehicle"].ownership
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_quotation_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVSelectPlanNewRenewalView(SelectPlanSupportView):
    def form_valid(self, form: forms.InsurancePlanForm):
        plan = form.cleaned_data["planes de seguro"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
                "plan_id": plan.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 7
        context["final_step"] = 9
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_premium_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateStepNewRenewalView(CreateIssuanceSupportView):
    def form_valid(self, form):
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        if premium.quotation_insurance_vehicle.expired:
            messages.warning(self.request, "Esta cotización ya expiró")
            return self.form_invalid(form)
        else:
            issuance = form.save(commit=False)
            issuance.issuance_type_id = 2
            issuance.consultant_registrar_id = self.kwargs["registrar_id"]
            issuance.consultant_seller_id = self.kwargs["seller_id"]
            issuance.insurance_plan_id = self.kwargs["plan_id"]

            raw_percentage = form.cleaned_data["plan_commission_percentage"]
            issuance.plan_commission_percentage = raw_percentage / 100
            seller = shortcuts.get_object_or_404(
                rrgg.models.Consultant, id=self.kwargs["seller_id"]
            )
            # TODO: comisión del vendedor por ahora solo tomara de referencia
            # la comisión actual registrada en el momento de la creación
            issuance.seller_commission_percentage = (
                seller.commission_rate.renewal
            )
            issuance.save()
            issuance.quotation_vehicle_premiums.add(premium)
            return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 8
        context["final_step"] = 9
        context["new_renewal_form"] = True
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self.kwargs["seller_id"]
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateStepNewRenewalView(UpdateIssuanceSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 8
        context["final_step"] = 9
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_plan_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.object.quotation_vehicle_premium.id,
                "seller_id": self.object.consultant_seller.id,
            },
        )
        return context


class IIVAddDocumentNewRenewalView(AddDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 9
        context["final_step"] = 9
        context["new_renewal_form"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_step_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDeleteDocumentNewRenewalView(DeleteDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 9
        context["final_step"] = 9
        context["return"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_nr",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


# ISSUANCE - RENEWAL 2


class IIVPremiumListRenewalStepView(PremiumListSupportView):
    template_name = (
        "rrggweb/issuance/insurance/vehicle/renewal/list_premium.html"
    )

    def _get_data(self):
        issuance = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        seller = issuance.consultant_seller
        premium = issuance.quotation_vehicle_premiums.first()
        customer = premium.quotation_insurance_vehicle.customer.pick
        return issuance, seller, customer

    def get_queryset(self):
        return self._get_data()[0].quotation_vehicle_premiums.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["seller"] = self._get_data()[1]
        context["customer"] = self._get_data()[2]
        context["new_register"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:search_vehicle_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": context["seller"].id,
                "customer_id": context["customer"].id,
            },
        )
        context["select_currency_ivr"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:select_currency_ivr_ns",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": context["seller"].id,
                "customer_id": context["customer"].id,
            },
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_renewal",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVCreateQuotationRenewalView(
    rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/form.html"
    model = rrgg.models.QuotationInsuranceVehicle
    fields = ["insured_amount", "currency"]

    def _get_data(self):
        previous_renewal = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        premium = previous_renewal.quotation_vehicle_premiums.first()
        quotation = premium.quotation_insurance_vehicle
        return previous_renewal, quotation

    def get_initial(self):
        initial = super().get_initial()
        initial["insured_amount"] = self._get_data()[1].insured_amount
        initial["currency"] = self._get_data()[1].currency
        return initial

    def form_valid(self, form):
        form.instance.risk_id = 1
        form.instance.consultant_registrar_id = self.kwargs["registrar_id"]
        form.instance.consultant_seller_id = self.kwargs["registrar_id"]
        form.instance.customer_id = self._get_data()[1].customer.id
        form.instance.vehicle_id = self._get_data()[1].vehicle.id
        form.instance.source = "renewal"
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["subtitle"] = "Crear registro vehicular"
        context["initial_step"] = 1
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self._get_data()[0].consultant_seller_id
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=self._get_data()[1].customer.id
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=self._get_data()[1].vehicle.id
        )
        context["owner"] = context["vehicle"].ownership

        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_renewal",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVUpdateQuotationStepRenewalView(UpdateQuotationSupportView):
    def _get_data(self):
        return shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_premium",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "quotation_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 1
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self._get_data().consultant_seller_id
        )
        context["customer"] = self.object.customer
        context["vehicle"] = self.object.vehicle
        context["owner"] = context["vehicle"].ownership
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:list_renewal",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class IIVCreatePremiumRenewalView(
    rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/quotation/insurance/vehicle/premium_form.html"
    model = rrgg.models.QuotationInsuranceVehiclePremium
    fields = ["insurance_vehicle_ratio", "amount", "rate"]

    def _get_data(self):
        issuance = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        premium = issuance.quotation_vehicle_premiums.first()
        ivr = premium.insurance_vehicle_ratio
        return issuance, ivr

    def get_initial(self):
        initial = super().get_initial()
        initial["insurance_vehicle_ratio"] = self._get_data()[1]
        return initial

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["insurance_vehicle_ratio"].required = False
        return form

    def form_valid(self, form):
        ivr = self._get_data()[1]
        form.instance.insurance_vehicle_ratio = ivr
        form.instance.tax_percentage = ivr.tax
        form.instance.emission_right_percentage = ivr.emission_right
        form.instance.quotation_insurance_vehicle_id = self.kwargs[
            "quotation_id"
        ]
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.object.id,
                "seller_id": self._get_data()[0].consultant_seller_id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["subtitle"] = "Registrar prima"
        context["initial_step"] = 2
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self._get_data()[0].consultant_seller_id
        )
        quotation = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehicle,
            id=self.kwargs["quotation_id"],
        )
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=quotation.customer.id
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=quotation.vehicle.id
        )
        context["owner"] = quotation.vehicle.ownership
        context["risk_selector"] = forms.RiskForm(risk_id=quotation.risk_id)
        context["insured_amount"] = quotation.insured_amount
        context["renewal_form"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_quotation_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "quotation_id": self.kwargs["quotation_id"],
            },
        )
        return context


class IIVUpdatePremiumStepRenewalView(UpdatePremiumSupportView):
    def _get_data(self):
        issuance = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        premium = issuance.quotation_vehicle_premiums.first()
        ivr = premium.insurance_vehicle_ratio
        return issuance, ivr

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["insurance_vehicle_ratio"].required = False
        return form

    def form_valid(self, form):
        form.instance.insurance_vehicle_ratio = self._get_data()[1]
        return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.object.id,
                "seller_id": self._get_data()[0].consultant_seller_id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 2
        context["final_step"] = 4
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant, id=self._get_data()[0].consultant_seller_id
        )
        quotation = self.object.quotation_insurance_vehicle
        context["customer"] = shortcuts.get_object_or_404(
            rrgg.models.CustomerMembership, id=quotation.customer.id
        )
        context["vehicle"] = shortcuts.get_object_or_404(
            rrgg.models.Vehicle, id=quotation.vehicle.id
        )
        context["owner"] = context["vehicle"].ownership
        context["risk_selector"] = forms.RiskForm(risk_id=quotation.risk_id)
        context["renewal_form"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_quotation_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "quotation_id": self.object.quotation_insurance_vehicle.id,
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVUpdateNaturalPersonRenewalView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateLegalPersonRenewalView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateVehicleRenewalView(UpdateVehicleSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVUpdateOwnerRenewalView(UpdateOwnerSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVChangeSellerRenewalView(LoginRequiredMixin, FormView):
    template_name = "rrggweb/quotation/insurance/vehicle/seller_form.html"
    form_class = forms.SellerForm

    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["subtitle"] = "Seleccionar asesor"
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class IIVCreateStepRenewalView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = "rrggweb/issuance/insurance/vehicle/form.html"
    model = rrgg.models.IssuanceInsuranceVehicle
    fields = [
        "policy",
        "collection_document",
        "issuance_date",
        "initial_validity",
        "final_validity",
        "plan_commission_percentage",
        "payment_method",
    ]

    def _get_data(self):
        return shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )

    def get_initial(self):
        initial = super().get_initial()
        plan = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self._get_data().insurance_plan.id
        )
        kcs_commission_percentage = round(plan.commission * 100, 1)
        initial["plan_commission_percentage"] = kcs_commission_percentage
        initial["policy"] = self._get_data().policy
        initial["initial_validity"] = self._get_data().initial_validity
        return initial

    def form_valid(self, form):
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        if premium.quotation_insurance_vehicle.expired:
            messages.warning(self.request, "Esta cotización ya expiró")
            return self.form_invalid(form)
        else:
            issuance = form.save(commit=False)
            issuance.issuance_type_id = 2
            issuance.insurance_plan_id = self._get_data().insurance_plan.id

            raw_percentage = form.cleaned_data["plan_commission_percentage"]
            issuance.plan_commission_percentage = raw_percentage / 100
            issuance.consultant_registrar_id = self.kwargs["registrar_id"]
            seller = shortcuts.get_object_or_404(
                rrgg.models.Consultant,
                id=self.kwargs["seller_id"],
            )
            issuance.consultant_seller_id = seller.id
            issuance.seller_commission_percentage = (
                seller.commission_rate.renewal
            )
            issuance.save()
            issuance.quotation_vehicle_premiums.add(premium)
            return super().form_valid(form)

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["subtitle"] = "Crear emisión"
        context["initial_step"] = 3
        context["final_step"] = 4
        context["renewal_form"] = True
        context["seller"] = shortcuts.get_object_or_404(
            rrgg.models.Consultant,
            id=self.kwargs["seller_id"],
        )
        context["change_seller"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:change_seller_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": context["seller"].id,
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
            },
        )
        context["premium"] = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        context["ratio"] = context["premium"].insurance_vehicle_ratio
        context["insurance_plan"] = shortcuts.get_object_or_404(
            rrgg.models.InsurancePlan, id=self._get_data().insurance_plan.id
        )
        context["quotation"] = context["premium"].quotation_insurance_vehicle
        context["customer"] = context["quotation"].customer.pick
        context["vehicle"] = context["quotation"].vehicle
        if isinstance(
            context["vehicle"].ownership.pick, rrgg.models.NaturalPerson
        ):
            context["owner"] = context["vehicle"].ownership.pick
            context["update_owner"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_owner",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "owner_id": context["owner"].id,
                    "issuance_id": self.kwargs["issuance_id"],
                    "premium_id": self.kwargs["premium_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
        premium = shortcuts.get_object_or_404(
            rrgg.models.QuotationInsuranceVehiclePremium,
            id=self.kwargs["premium_id"],
        )
        if isinstance(context["customer"], rrgg.models.NaturalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_natural_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "natural_person_id": context["customer"].id,
                    "issuance_id": self.kwargs["issuance_id"],
                    "premium_id": self.kwargs["premium_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
        elif isinstance(context["customer"], rrgg.models.LegalPerson):
            context["update_customer"] = urls.reverse(
                "rrggweb:issuance:insurance:vehicle:update_legal_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "legal_person_id": context["customer"].id,
                    "issuance_id": self.kwargs["issuance_id"],
                    "premium_id": self.kwargs["premium_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
        else:
            pass
        context["update_vehicle"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_vehicle",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "vehicle_id": premium.quotation_insurance_vehicle.vehicle.id,
                "issuance_id": self.kwargs["issuance_id"],
                "premium_id": self.kwargs["premium_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_premium_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.kwargs["premium_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVUpdateIssuanceStepRenewalView(UpdateIssuanceSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.object.id,
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 5
        context["renewal_form"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_premium_step",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "premium_id": self.object.quotation_vehicle_premium.id,
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVCreateDocumentRenewalView(AddDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 4
        context["final_step"] = 4
        context["renewal_form"] = True
        context["previous_page"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:update_step_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


class IIVDeleteDocumentRenewalView(DeleteDocumentSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "EMISIÓN VEHICULAR - RENOVACIÓN"
        context["initial_step"] = 5
        context["final_step"] = 5
        context["return"] = urls.reverse(
            "rrggweb:issuance:insurance:vehicle:create_document_r",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "issuance_id": self.kwargs["issuance_id"],
            },
        )
        return context


# ------------------------------

# COLLECTION VIEW


class CollectionView(TemplateView):
    template_name = "rrggweb/collection.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["seguros"] = [
            SeguroItem(
                "Seguro de vehicular",
                urls.reverse(
                    "rrggweb:collection:insurance:vehicle:list",
                    kwargs={"registrar_id": self.kwargs["registrar_id"]},
                ),
            ),
            SeguroItem("Seguro de vida", ""),
            SeguroItem("Seguro de accidentes", ""),
            SeguroItem("Seguro de salud", ""),
        ]
        return context


class CollectionInsuranceVehicleListView(ListView):
    template_name = "rrggweb/collection/insurance/vehicle/list.html"
    model = rrgg.models.CollectionInsuranceVehicle


class CollectionInsuranceVehicleCreateCollectionView(
    rrgg_mixins.RrggBootstrapDisplayMixin, CreateView
):
    template_name = (
        "rrggweb/collection/insurance/vehicle/create_collection.html"
    )
    model = rrgg.models.CollectionInsuranceVehicle
    fields = [
        "expiration_date",
        "payment_date",
        "payment_receipt",
        "issue",
        "amount",
    ]

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields["payment_date"].required = False
        form.fields["payment_receipt"].required = False
        return form

    def form_valid(self, form):
        form.instance.issuance_vehicle_id = self.kwargs["issuance_id"]
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["issuance"] = shortcuts.get_object_or_404(
            rrgg.models.IssuanceInsuranceVehicle,
            id=self.kwargs["issuance_id"],
        )
        return context

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:collection:insurance:vehicle:list",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
            },
        )


# ------------------------------

# CLIENT MAIN


class CustomerMembershipListView(ListView):
    template_name = "rrggweb/client/list.html"
    model = rrgg.models.CustomerMembership
    context_object_name = "memberships"
    paginate_by = 10
    ordering = ["-id"]

    def get_queryset(self):
        query = self.request.GET.get("q")
        if query:
            natural_persons = rrgg.models.NaturalPerson.objects.filter(
                Q(given_name__icontains=query)
                | Q(first_surname__icontains=query)
                | Q(second_surname__icontains=query)
                | Q(birthdate__icontains=query)
                | Q(document_number__icontains=query)
            ).distinct()
            legal_persons = rrgg.models.LegalPerson.objects.filter(
                Q(registered_name__icontains=query)
                | Q(general_manager__icontains=query)
                | Q(anniversary_date__icontains=query)
                | Q(document_number__icontains=query)
            ).distinct()
            return list(natural_persons) + list(legal_persons)
        else:
            return super().get_queryset()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Lista de clientes"
        context["register"] = urls.reverse(
            "rrggweb:customer_membership:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        context["search_query"] = self.request.GET.get("q", "")
        context["num_results"] = len(context["memberships"])
        context["num_registers"] = self.model.objects.count()
        if self.request.GET.get("page"):
            context["page_number"] = int(self.request.GET.get("page"))
        return context


class CustomerMembershipDetailView(
    LoginRequiredMixin, rrgg_mixins.RrggBootstrapDisplayMixin, DetailView
):
    template_name = "rrggweb/client/detail.html"
    model = rrgg.models.CustomerMembership
    fields = "__all__"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Detalle de cliente"
        return_url = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        if self.request.GET.get("page"):
            page_number = self.request.GET.get("page")
            context["previous_page"] = f"{return_url}?page={page_number}"
        else:
            context["previous_page"] = return_url
        return context


class CMUpdateNaturalPersonView(UpdateNaturalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Editar cliente"
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMUpdateLegalPersonView(UpdateLegalPersonSupportView):
    def get_success_url(self):
        return urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Editar cliente"
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMChangeAssociateConsultantView(
    rrgg_mixins.RrggBootstrapDisplayMixin, UpdateView
):
    template_name = "rrggweb/client/form.html"
    model = rrgg.models.CustomerMembership
    fields = ["seller"]
    pk_url_kwarg = "customer_id"

    def get_success_url(self):
        success_url = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        if self.request.GET.get("page"):
            page_number = self.request.GET.get("page")
            return f"{success_url}?page={page_number}"
        else:
            return success_url

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Cambiar asesor"
        return_url = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        if self.request.GET.get("page"):
            page_number = self.request.GET.get("page")
            context["previous_page"] = f"{return_url}?page={page_number}"
        else:
            context["previous_page"] = return_url
        return context


class CMDeletePersonSupportView(DeleteView):
    template_name = "rrggweb/client/delete_form.html"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            self.object.membership.delete()
            self.object.delete()
            return redirect(self.get_success_url())

        except ProtectedError:
            messages.error(
                request,
                (
                    "No se puede eliminar este cliente porque tiene"
                    " registros vehiculares asociados."
                ),
            )
            return redirect(request.get_full_path())

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["subtitle"] = "Eliminar cliente"
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMDeleteNaturalPersonView(CMDeletePersonSupportView):
    model = rrgg.models.NaturalPerson


class CMDeleteLegalPersonView(CMDeletePersonSupportView):
    model = rrgg.models.LegalPerson


# CLIENT - STEPS


class CMSelectRoleFormView(SelectRoleSupportView):
    def form_valid(self, form: forms.RoleForm):
        role = form.cleaned_data["roles"]
        self.success_url = urls.reverse(
            "rrggweb:customer_membership:select_seller",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "role_id": role.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["initial_step"] = 1
        context["final_step"] = 3
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMSelectSellerFormView(SelectSellerSupportView):
    def form_valid(self, form: forms.SellerForm):
        seller = form.cleaned_data["asesores"]
        self.success_url = urls.reverse(
            "rrggweb:customer_membership:select_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": seller.id,
            },
        )
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["initial_step"] = 1
        context["final_step"] = 3
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMSelectCustomerFormView(SelectCustomerSupportView):
    def form_valid(self, form: forms.SelectCustomerForm):
        type_customer = form.cleaned_data["type_customer"]
        if type_customer == "persona_natural":
            create_customer_url = urls.reverse(
                "rrggweb:customer_membership:create_natural_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = create_customer_url
        elif type_customer == "persona_jurídica":
            create_customer_url = urls.reverse(
                "rrggweb:customer_membership:create_legal_person",
                kwargs={
                    "registrar_id": self.kwargs["registrar_id"],
                    "seller_id": self.kwargs["seller_id"],
                },
            )
            self.success_url = create_customer_url
        else:
            pass
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["initial_step"] = 2
        context["final_step"] = 3
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:select_role",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class CMCreateNaturalPersonView(CreateNaturalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            natural_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["initial_step"] = 3
        context["final_step"] = 3
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:select_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


class CMCreateLegalPersonView(CreateLegalPersonSupportView):
    def get_success_url(self):
        rrgg.models.CustomerMembership.objects.create(
            legal_person=self.object,
            seller_id=self.kwargs["seller_id"],
        )
        return urls.reverse(
            "rrggweb:customer_membership:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "CLIENTES"
        context["initial_step"] = 3
        context["final_step"] = 3
        context["previous_page"] = urls.reverse(
            "rrggweb:customer_membership:select_customer",
            kwargs={
                "registrar_id": self.kwargs["registrar_id"],
                "seller_id": self.kwargs["seller_id"],
            },
        )
        return context


# View para la data historica


class HistoricalDataListView(ListView):
    model = rrgg.models.HistoricalData
    template_name = "rrggweb/historical_data/list.html"
    context_object_name = "data"


class HistoricalDataDetailView(DetailView):
    model = rrgg.models.HistoricalData
    template_name = "rrggweb/historical_data/detail.html"
    context_object_name = "record"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["previous_page"] = urls.reverse(
            "rrggweb:historical_data:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context


class HistoricalDataUpdateView(UpdateView):
    model = rrgg.models.HistoricalData
    template_name = "rrggweb/historical_data/form.html"
    pk_url_kwarg = "historical_data_id"
    fields = "__all__"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["previous_page"] = urls.reverse(
            "rrggweb:historical_data:list",
            kwargs={"registrar_id": self.kwargs["registrar_id"]},
        )
        return context

    def get_success_url(self):
        return urls.reverse(
            "rrggweb:historical_data:detail",
            kwargs={"pk": self.object.id},
        )
